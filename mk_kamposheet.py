# PRG1 ライブラリ準備
import streamlit as st
import pandas as pd
import os
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

# PRG2 関数準備
# 通常のセルのデザイン
def cell_design(cell):
    side_normal = Side(style='thin', color='000000')
    border = Border(top=side_normal, bottom=side_normal, left=side_normal, right=side_normal)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name='ＭＳ Ｐゴシック', size=11, color='000000')

# 右端のセル全体折り返し表示追加
def cell_design_right(cell):
    side_normal = Side(style='thin', color='000000')
    border = Border(top=side_normal, bottom=side_normal, left=side_normal, right=side_normal)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(name='ＭＳ Ｐゴシック', size=11, color='000000')

# 表下端のセル外線下を中間太さで塗る
def cell_design_bottom(cell):
    side_normal = Side(style='thin', color='000000')
    side_bottom = Side(style='medium', color='000000')
    border = Border(top=side_normal, bottom=side_bottom, left=side_normal, right=side_normal)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name='ＭＳ Ｐゴシック', size=11, color='000000')

# 表右下の外線下を中間太さ＋セル全体折り返し表示追加
def cell_design_rightbottom(cell):
    side_normal = Side(style='thin', color='000000')
    side_bottom = Side(style='medium', color='000000')
    border = Border(top=side_normal, bottom=side_bottom, left=side_normal, right=side_normal)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(name='ＭＳ Ｐゴシック', size=11, color='000000')

# csv読み込み
def read_kampocsv():
    shouyaku_dataname = 'shouyaku_list'
    shouyaku_csv = shouyaku_dataname + '.csv'
    shouyaku_path = os.path.join(os.getcwd(), shouyaku_csv)
    df_shouyaku = pd.read_csv(shouyaku_path, header=0)

    kampo_dataname = 'kampo_list'
    kampo_csv = kampo_dataname + '.csv'
    kampo_path = os.path.join(os.getcwd(), kampo_csv)
    df_kampo = pd.read_csv(kampo_path, header=0)
    return df_shouyaku, df_kampo


# 選択漢方の重複を取り除くリストを作成
def drop_dupllicated_kampo(df_kampo, df_shouyaku, kampo_select, shouyaku_select):
    # PRG1 漢方の抽出
    select_kampo = df_kampo[df_kampo['検索用'].isin(kampo_select)].loc[:, ['漢方名', '配合生薬']]
    # PRG2 構成生薬と数の抽出
    kampo_name = select_kampo['漢方名'].tolist()
    kampo_components = select_kampo['配合生薬'].tolist()
    # PRG2 内訳確認用
    result_list = []
    # PRG3 一つのリストに内包する
    pre_list = []
    for name, components in zip(kampo_name, kampo_components):
        one_shouyaku = components.split('、')
        result = f"{name}の生薬数：{len(one_shouyaku)}"
        result_list.append(result)
        pre_list += one_shouyaku
        
    # PRG4 単独で選んだ生薬を追加する
    if len(shouyaku_select) > 0:
        select_shouyaku = df_shouyaku[df_shouyaku['検索用'].isin(shouyaku_select)]
        pre_list += select_shouyaku['配合生薬'].tolist()

    # PRG5 重複削除
    all_shouyaku = list(set(pre_list))
    # 重複生薬を確認できるようにする
    if len(all_shouyaku) < len(pre_list):
        dup_name = [x for x in set(pre_list) if pre_list.count(x) > 1]
        dup_text = ('、').join(dup_name)
    else:
        dup_name = []
        dup_text = 'ありません。'

    # PRG6 生薬CSVから生薬の情報を抽出する
    df_candidate = df_shouyaku[df_shouyaku['配合生薬'].isin(all_shouyaku)]

    # PRG7 個別に取り除きたい候補をリスト化する
    drop_candidate = df_candidate['検索用'].tolist()
    return result_list, dup_name, dup_text, df_candidate, drop_candidate

# Excel出力用の生薬を出力する
def extract_herballist(df_candidate, dup_name, dup_text, drop_select):
    #全生薬を除去選択した場合、そもそも漢方薬や生薬リストが未入力
    if len(drop_select) == len(df_candidate) or len(df_candidate) == 0:
        none_text = '構成生薬がありません。  \nシート作成できません！'
        return none_text, dup_name, dup_text
    elif len(drop_select) == 0:
        return df_candidate, dup_name, dup_text
    else:
        df_select = df_candidate[~df_candidate['検索用'].isin(drop_select)]
        _drop_select = set([x.split('(')[0] for x in drop_select]) & set(dup_name)
        dup_name = list(set(dup_name) - _drop_select)
        if len(dup_name) == 0:
            dup_text = 'ありません。'
        else:
            dup_text = ('、').join(dup_name)
        return df_select, dup_name, dup_text

# Excelシート作成num_rowcomment
def mk_kamposheet(df_select):
    # PRG1 作成するシートの枚数を計算する
    num_rowcomment = 6
    if len(df_select) % num_rowcomment == 0:
        sheet_num = len(df_select) // num_rowcomment
    else:
        sheet_num = len(df_select) // num_rowcomment + 1

    # PRG2 excelのインスタンス化
    wb_shouyaku = xl.Workbook()

    # PRG3 シート作成 シートが作成されてなければ
    sheet_name = '薬情'
    if not sheet_name in wb_shouyaku.sheetnames:
        wb_shouyaku.create_sheet(title=sheet_name)
    remove_name = 'Sheet'
    if remove_name in wb_shouyaku.sheetnames:
        wb_shouyaku.remove(wb_shouyaku[remove_name])

    # PRG4 書き込むシートを変数作成
    ws = wb_shouyaku[sheet_name]

    # PRG5 中身を作成
    # PRG5_0 列幅
    ws.column_dimensions['A'].width=6
    ws.column_dimensions['B'].width=18
    ws.column_dimensions['C'].width=11
    ws.column_dimensions['D'].width=17
    ws.column_dimensions['E'].width=11
    ws.column_dimensions['F'].width=63
    ws.column_dimensions['G'].width=15

    for i in range(sheet_num):
    # PRG5_1 タイトル書き込み位置
        num_rowsheet = 2 + num_rowcomment
        title_iloc = 1 + num_rowsheet*i
        
    # PRG5_2 タイトル、確定事項書き込み
        # PRG5_2_1 タイトル書き込み
        title_list = ['注意', '写真', '配合生薬', '効能', '薬味', '解説']
        for num, title_name in enumerate(title_list):
            # 行の高さ
            ws.row_dimensions[title_iloc].height = 18
            # 内容
            ws.cell(row=title_iloc, column=num+1).value = title_name
            # 書式設定
            cell_design(ws.cell(row=title_iloc, column=num+1))
        
        # PRG5_2_2 特記事項
        # 行の高さ
        bottom_row_number = title_iloc + num_rowcomment + 1
        ws.row_dimensions[bottom_row_number].height = 18
        # 内容
        ws.cell(row=bottom_row_number, column=1).value = '特記事項：'
        # 書式設定
        ws.cell(row=bottom_row_number, column=1).alignment = Alignment(vertical='center')
        ws.cell(row=bottom_row_number, column=1).font = Font(name='ＭＳ Ｐゴシック', size=11, color='000000')
        
        # PRG5_2_3 行の高さ
        for cell_num in range(num_rowcomment):
            height_iloc = title_iloc + cell_num + 1
            row_height = 80
            ws.row_dimensions[height_iloc].height = row_height
        
        # PRG5_4 改ページ (行) 個別指定
        row_break = Break(bottom_row_number)
        ws.row_breaks.append(row_break)

    # PRG6 生薬データ内容の書き込み
        df_part = df_select.iloc[i*num_rowcomment:i*num_rowcomment+num_rowcomment, :].reset_index(drop=True)
        for row_num, cols in df_part.iterrows():
            # 入力行
            input_row = title_iloc + row_num + 1
            for col_num, col in enumerate(cols):
                if col_num == 0:
                    # PRG6_3_1 生薬名
                    ws.cell(row=input_row, column=3).value = col
                elif col_num == 2:
                    # PRG6_3_1 写真
                    # 生薬名の写真取得
                    picture_name = col
                    picture_file = picture_name + '.jpg'
                    # 写真のパスを取得
                    picture_path = os.path.join(os.getcwd(), 'shouyaku_images', picture_file)
                    # 写真の貼り付け引数を取得
                    img_to_excel = Image(picture_path)
                    img_position = ws.cell(row=input_row, column=col_num).coordinate
                    ws.add_image(img_to_excel, img_position)
                elif col_num >= 4:
                    # PRG6_3_3 効能, 薬味, 解説
                    ws.cell(row=input_row, column=col_num).value = col
                    
                if row_num == num_rowcomment-1 and col_num == 6:
                    cell_design_rightbottom(ws.cell(row=input_row, column=col_num))
                elif row_num < num_rowcomment-1 and col_num == 6:
                    cell_design_right(ws.cell(row=input_row, column=col_num))
                elif row_num == num_rowcomment-1:
                    cell_design_bottom(ws.cell(row=input_row, column=col_num+1))
                else:
                    cell_design(ws.cell(row=input_row, column=col_num+1))
            

    # PRG7 ページレイアウトの設定
    # 印刷の向き
    # 用紙サイズ・印刷の向き
    ws.set_printer_settings(
        Worksheet.PAPERSIZE_A4,
        Worksheet.ORIENTATION_LANDSCAPE
    )

    CM = 1 / 2.54
    # 余白
    ws.page_margins.top = 1.9 * CM
    ws.page_margins.right = 1.8 * CM
    ws.page_margins.bottom = 1.9 * CM
    ws.page_margins.left = 1.8 * CM
    # 余白 (ヘッダー・フッター)
    ws.page_margins.header = 0.8 * CM
    ws.page_margins.footer = 0.8 * CM
    # ヘッダー (全ページ共通)
    # 中央
    ws.oddHeader.center.text = '漢方薬　構成生薬'
    ws.oddHeader.center.font = 'HGS行書体, Bold'
    ws.oddHeader.center.size = 20
    # 右
    ws.oddHeader.right.text = '無断複写・転用を禁じます'
    ws.oddHeader.right.font = 'ＭＳ Ｐゴシック'
    ws.oddHeader.right.size = 11
    # フッダー (全ページ共通)
    ws.oddFooter.right.text = '&P/&N'
    ws.oddFooter.right.font = 'Arial Unicode MS, Bold'
    ws.oddFooter.right.size = 11
    # ドキュメントに合わせて配置
    ws.scaleWithDoc = True
    # ページ余白に合わせて配置
    ws.alignWithMargins = True
    # ビュー設定
    ws.sheet_view.view = 'pageLayout'

    # PRG8 保存する
    save_name = 'streamlit_test'
    save_file = save_name + '.xlsx'
    save_path = os.path.join(os.getcwd(), save_file)
    save_excel = wb_shouyaku.save(save_path)
    return save_path


# PRG3 streamlitの表記を作成する
# CSV読み込み
df_shouyaku, df_kampo = read_kampocsv()
# セッション情報の初期化
if "form_flag_1" not in st.session_state:
    st.session_state.form_flag_1 = False

if "form_flag_2" not in st.session_state:
    st.session_state.form_flag_2 = False

def change_flag_1():
    st.session_state.form_flag_1 = True
    st.session_state.form_flag_2 = False

def change_flag_2():
    st.session_state.form_flag_2 = True

st.title('構成生薬シート作成　アプリ')
st.header('どんなアプリ？')
st.write('取捨選択した漢方・生薬末に含まれる\n  \n:red[生薬情報]を**重複なく**EXCELシートに:red[作成]します')
with st.form(key="form_1"):
    st.subheader('STEP1:　使用する漢方を選択する')
    multi_select_1 = st.multiselect(
        'ボックス内で:blue[複数]選択することも可能です',
        df_kampo['検索用'],
    []
    )
    st.caption('ひらがな入力するとサジェストで探しやすくなります')
    st.subheader('STEP2:　使用する生薬・生薬末を選択する')
    multi_select_2 = st.multiselect(
        'ボックス内で:blue[複数]選択することも可能です',
        df_shouyaku['検索用'],
        []
    )
    st.caption('ひらがな入力するとサジェストで探しやすくなります')
    st.subheader('STEP3:　「生薬抽出」ボタンを押す')
    st.write(':red[選びなおした]場合、:red[再度クリック]する必要あり。  \n  \n押さない場合、結果は反映されないのでご注意を！')
    select_btn_1 = st.form_submit_button(label="生薬抽出", on_click=change_flag_1)

if st.session_state.form_flag_1:
    result_list, dup_name_1, dup_text_1, df_candidate, drop_candidate = drop_dupllicated_kampo(df_kampo, df_shouyaku, multi_select_1, multi_select_2)

    
    
    st.subheader('STEP4:　取り除きたい生薬・生薬末を選択する')
    st.write('必要ない場合はSTEP5へ！')
    drop_select = st.multiselect(
        'ボックス内で:blue[複数]選択することも可能です',
        drop_candidate,
        []
    )
    st.caption('ひらがな入力するとサジェストで探しやすくなります')
    df_select, dup_name_2, dup_text_2 = extract_herballist(df_candidate, dup_name_1, dup_text_1, drop_select)
    # 未入力時や全構成生薬を除去したときに確認OKボタンが現れないようにする
    if type(df_select) is str:
        st.write(df_select)
        st.session_state.form_flag_2 = False
    else:
        # 除去されず選択生薬が残っているときのみ表示される
        if len(df_select) > 0:
            st.subheader('STEP5 問題なければ「確認OK」ボタン')
            select_btn_2 = st.button(label="確認OK", on_click=change_flag_2)
        # 内訳を表記
        st.subheader('簡易　確認用')
        st.write(f'**重複なし生薬総数：{len(df_select)}**')
        st.write(("、").join(df_select["配合生薬"].tolist()))
        st.write(f'**除去選択前の重複生薬{len(dup_name_1)}種類**')
        st.write(dup_text_1)
        st.write('**取捨選択した漢方・生薬の数など**')
        st.write('  \n'.join(result_list))
        st.write(f'選択した生薬・生薬末数：{len(multi_select_2)}')
        st.write(f'除去した生薬・生薬末数：{len(drop_select)}  \n')

        st.write(':red[「確認OK」ボタン押すと下に「Download」ボタンが出現します]')

# ダウンロードボタンが表示される
if st.session_state.form_flag_2:
    st.subheader('STEP6 「Download」ボタンからExcelをダウンロード')
    save_path = mk_kamposheet(df_select)
    ASIA_TOKYO = ZoneInfo('Asia/Tokyo')
    dt_now = datetime.now(ASIA_TOKYO).strftime('%Y年%m月%d日%H時%M分%S秒')
    with open(save_path, "rb") as d:
        st.download_button("Download", d, file_name=f"構成生薬_{dt_now}.xlsx")
    
    st.write('ファイル名は、「:blue[構成生薬_年月日時分秒.xlsx]」です')
    st.write('Excelを開いて:red[編集を有効にする]をクリックするとページが整います。')

        






